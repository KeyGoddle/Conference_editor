from __future__ import annotations

import argparse
import json
import re
import sys
from dataclasses import dataclass
from functools import lru_cache
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except ImportError as exc:  # pragma: no cover - helpful runtime message
    raise SystemExit(
        "Не найден пакет openpyxl. Установи зависимости: python3 -m pip install -r requirements.txt"
    ) from exc

try:
    from PIL import Image, ImageDraw, ImageFont
except ImportError as exc:  # pragma: no cover - helpful runtime message
    raise SystemExit(
        "Не найден пакет Pillow. Установи зависимости: python3 -m pip install -r requirements.txt"
    ) from exc


FieldMap = dict[str, str]


HEADER_ALIASES: dict[str, list[str]] = {
    "full_name_ru": [
        "фио на русском",
        "фио рус",
        "фио ru",
        "фио",
    ],
    "full_name_en": [
        "фио на англе",
        "фио на английском",
        "фио англ",
        "фио en",
    ],
    "city_ru": [
        "город на русском",
        "город рус",
        "город ru",
        "город",
    ],
    "city_en": [
        "город на англе",
        "город на английском",
        "город англ",
        "город en",
    ],
    "talk_title": [
        "название доклада",
        "доклад",
        "тема доклада",
        "тема",
    ],
}


DEFAULT_FONT_CANDIDATES = [
    "/System/Library/Fonts/Supplemental/Arial.ttf",
    "/System/Library/Fonts/Supplemental/Arial Bold.ttf",
    "/System/Library/Fonts/Supplemental/Arial Unicode.ttf",
    "/System/Library/Fonts/Supplemental/Helvetica.ttf",
    "/Library/Fonts/Arial.ttf",
    "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
    "/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf",
    "C:/Windows/Fonts/arial.ttf",
]

FONT_SEARCH_DIRS = [
    Path("/System/Library/Fonts"),
    Path("/System/Library/Fonts/Supplemental"),
    Path("/Library/Fonts"),
    Path.home() / "Library" / "Fonts",
    Path("/usr/share/fonts"),
    Path("/usr/local/share/fonts"),
    Path("C:/Windows/Fonts"),
]

FONT_EXTENSIONS = {".ttf", ".otf", ".ttc"}

A4_PORTRAIT = (2480, 3508)
A4_LANDSCAPE = (3508, 2480)
PDF_DPI = 300
BADGE_PDF_MARGIN = 90
BADGE_PDF_GAP = 40
BADGE_MIN_SCALE = 0.65


DEFAULT_CONFIG: dict[str, Any] = {
    "badges": {
        "output_prefix": "badge",
        "side_gap": 24,
        "sides": [
            {
                "name": "ru",
                "blocks": [
                    {
                        "field": "full_name_ru",
                        "x": 500,
                        "y": 450,
                        "max_width": 760,
                        "max_height": 130,
                        "font_size": 50,
                        "fill": "#111111",
                    },
                    {
                        "field": "city_ru",
                        "x": 500,
                        "y": 635,
                        "max_width": 700,
                        "max_height": 90,
                        "font_size": 34,
                        "fill": "#111111",
                    },
                ],
            },
            {
                "name": "en",
                "blocks": [
                    {
                        "field": "full_name_en",
                        "x": 500,
                        "y": 450,
                        "max_width": 760,
                        "max_height": 130,
                        "font_size": 50,
                        "fill": "#111111",
                    },
                    {
                        "field": "city_en",
                        "x": 500,
                        "y": 635,
                        "max_width": 700,
                        "max_height": 90,
                        "font_size": 34,
                        "fill": "#111111",
                    },
                ],
            },
        ],
    },
    "certificates": {
        "output_prefix": "certificate",
        "blocks": [
            {
                "field": "full_name_ru",
                "x": 875,
                "y": 620,
                "max_width": 1200,
                "max_height": 130,
                "font_size": 58,
                "fill": "#111111",
            },
            {
                "field": "talk_title",
                "x": 875,
                "y": 820,
                "max_width": 1250,
                "max_height": 180,
                "font_size": 38,
                "fill": "#111111",
            },
        ],
    },
}


@dataclass(frozen=True)
class TextBlock:
    field: str
    x: int
    y: int
    max_width: int
    max_height: int
    font_size: int
    fill: str = "#111111"
    font_path: str | None = None
    align: str = "center"
    bold: bool = False
    stroke_width: int = 0
    stroke_fill: str | None = None


@dataclass(frozen=True)
class BadgeSide:
    name: str
    blocks: list[TextBlock]


def normalize_header(value: Any) -> str:
    text = "" if value is None else str(value)
    text = text.strip().lower().replace("ё", "е")
    text = re.sub(r"\s+", " ", text)
    return text


def canonical_aliases(field: str) -> set[str]:
    aliases = HEADER_ALIASES.get(field, [field])
    return {normalize_header(alias) for alias in aliases}


def normalize_font_name(value: str) -> str:
    text = Path(value).stem if Path(value).suffix.lower() in FONT_EXTENSIONS else value
    text = text.strip().lower().replace("_", " ").replace("-", " ")
    text = re.sub(r"\s+", " ", text)
    return text


@lru_cache(maxsize=1)
def system_font_catalog() -> dict[str, str]:
    catalog: dict[str, str] = {}

    for directory in FONT_SEARCH_DIRS:
        if not directory.exists():
            continue
        for path in directory.rglob("*"):
            if path.suffix.lower() not in FONT_EXTENSIONS:
                continue

            name = normalize_font_name(path.stem)
            catalog.setdefault(name, str(path))
            for suffix in (" regular", " roman", " normal"):
                if name.endswith(suffix):
                    catalog.setdefault(name[: -len(suffix)], str(path))

    return catalog


def find_font_path(explicit_path: str | None = None, bold: bool = False) -> str:
    if explicit_path:
        path = Path(explicit_path).expanduser()
        if path.exists():
            return str(path)

        catalog = system_font_catalog()
        name = normalize_font_name(explicit_path)
        if bold:
            for bold_name in (f"{name} bold", f"{name} bd", name.replace(" regular", " bold")):
                if bold_name in catalog:
                    return catalog[bold_name]
        if name in catalog:
            return catalog[name]

        matches = [
            font_path
            for font_name, font_path in catalog.items()
            if name in font_name and ("bold" in font_name) == bold
        ]
        if matches:
            return matches[0]

        raise FileNotFoundError(f"Шрифт не найден: {explicit_path}")

    for candidate in DEFAULT_FONT_CANDIDATES:
        if bold and "Bold" not in candidate and "bold" not in candidate:
            continue
        if Path(candidate).exists():
            return candidate

    if bold:
        return find_font_path(None, bold=False)

    raise FileNotFoundError(
        "Не удалось найти системный шрифт с кириллицей. "
        "Укажи путь к .ttf/.otf в layout.json через поле font_path."
    )


def load_json_config(path: Path | None) -> dict[str, Any]:
    if path is None:
        return DEFAULT_CONFIG

    with path.open("r", encoding="utf-8") as file:
        user_config = json.load(file)

    config = json.loads(json.dumps(DEFAULT_CONFIG))
    for mode, mode_config in user_config.items():
        if isinstance(mode_config, dict):
            config.setdefault(mode, {})
            config[mode].update(mode_config)
    return config


def build_blocks(raw_blocks: list[dict[str, Any]], context: str) -> list[TextBlock]:
    blocks: list[TextBlock] = []
    for index, block in enumerate(raw_blocks, start=1):
        try:
            blocks.append(
                TextBlock(
                    field=str(block["field"]),
                    x=int(block["x"]),
                    y=int(block["y"]),
                    max_width=int(block["max_width"]),
                    max_height=int(block["max_height"]),
                    font_size=int(block["font_size"]),
                    fill=str(block.get("fill", "#111111")),
                    font_path=block.get("font_path"),
                    align=str(block.get("align", "center")),
                    bold=bool(block.get("bold", False)),
                    stroke_width=int(block.get("stroke_width", 0)),
                    stroke_fill=block.get("stroke_fill"),
                )
            )
        except KeyError as exc:
            raise ValueError(
                f"В блоке #{index} ({context}) не хватает поля {exc.args[0]!r}"
            ) from exc
    return blocks


def build_text_blocks(config: dict[str, Any], mode: str) -> list[TextBlock]:
    try:
        raw_blocks = config[mode]["blocks"]
    except KeyError as exc:
        raise ValueError(f"В настройках нет режима {mode!r} или списка blocks") from exc

    return build_blocks(raw_blocks, mode)


def build_badge_sides(config: dict[str, Any]) -> list[BadgeSide]:
    badge_config = config.get("badges")
    if not isinstance(badge_config, dict):
        raise ValueError("В настройках нет режима 'badges'")

    raw_sides = badge_config.get("sides")
    if raw_sides is None:
        # Backward compatibility for the old one-sided badge layout.
        return [BadgeSide(name="badge", blocks=build_text_blocks(config, "badges"))]

    sides: list[BadgeSide] = []
    for index, side in enumerate(raw_sides, start=1):
        name = str(side.get("name", f"side_{index}"))
        raw_blocks = side.get("blocks")
        if not isinstance(raw_blocks, list):
            raise ValueError(f"У стороны бейджа {name!r} нет списка blocks")
        sides.append(BadgeSide(name=name, blocks=build_blocks(raw_blocks, f"badges.{name}")))

    if not sides:
        raise ValueError("В badges.sides должен быть хотя бы один элемент")
    return sides


def read_excel_rows(excel_path: Path, fields: list[str], sheet_name: str | None) -> list[FieldMap]:
    workbook = load_workbook(excel_path, read_only=True, data_only=True)
    worksheet = workbook[sheet_name] if sheet_name else workbook.active

    rows = worksheet.iter_rows(values_only=True)
    try:
        raw_headers = next(rows)
    except StopIteration:
        return []

    normalized_headers = [normalize_header(header) for header in raw_headers]
    header_indexes: dict[str, int] = {}
    missing_fields: list[str] = []

    for field in fields:
        aliases = canonical_aliases(field)
        index = next(
            (i for i, header in enumerate(normalized_headers) if header in aliases),
            None,
        )
        if index is None:
            missing_fields.append(field)
        else:
            header_indexes[field] = index

    if missing_fields:
        readable_headers = ", ".join(str(header) for header in raw_headers if header)
        missing = ", ".join(missing_fields)
        raise ValueError(
            f"В файле {excel_path.name} не найдены колонки: {missing}. "
            f"Найденные заголовки: {readable_headers or 'нет заголовков'}"
        )

    result: list[FieldMap] = []
    for row in rows:
        item: FieldMap = {}
        for field, index in header_indexes.items():
            value = row[index] if index < len(row) else ""
            item[field] = "" if value is None else str(value).strip()
        if any(item.values()):
            result.append(item)
    return result


def text_size(draw: ImageDraw.ImageDraw, text: str, font: ImageFont.FreeTypeFont) -> tuple[int, int]:
    if not text:
        return 0, 0
    bbox = draw.multiline_textbbox((0, 0), text, font=font, spacing=8, align="center")
    return bbox[2] - bbox[0], bbox[3] - bbox[1]


def block_stroke_width(block: TextBlock) -> int:
    if block.stroke_width > 0:
        return block.stroke_width
    if block.bold:
        return max(1, block.font_size // 30)
    return 0


def wrap_text(draw: ImageDraw.ImageDraw, text: str, font: ImageFont.FreeTypeFont, max_width: int) -> str:
    paragraphs = str(text).splitlines() or [str(text)]
    wrapped_lines: list[str] = []

    for paragraph in paragraphs:
        words = paragraph.split()
        if not words:
            wrapped_lines.append("")
            continue

        line = words[0]
        for word in words[1:]:
            candidate = f"{line} {word}"
            width, _ = text_size(draw, candidate, font)
            if width <= max_width:
                line = candidate
            else:
                wrapped_lines.append(line)
                line = word
        wrapped_lines.append(line)

    return "\n".join(wrapped_lines)


def fit_text(
    draw: ImageDraw.ImageDraw,
    text: str,
    font_path: str,
    font_size: int,
    max_width: int,
    max_height: int,
) -> tuple[str, ImageFont.FreeTypeFont, int, int]:
    size = font_size
    while size >= 8:
        font = ImageFont.truetype(font_path, size=size)
        wrapped = wrap_text(draw, text, font, max_width)
        width, height = text_size(draw, wrapped, font)
        if width <= max_width and height <= max_height:
            return wrapped, font, width, height
        size -= 1

    font = ImageFont.truetype(font_path, size=8)
    wrapped = wrap_text(draw, text, font, max_width)
    width, height = text_size(draw, wrapped, font)
    return wrapped, font, width, height


def draw_centered_text(image: Image.Image, block: TextBlock, text: str) -> None:
    if not text:
        return

    draw = ImageDraw.Draw(image)
    font_path = find_font_path(block.font_path, bold=block.bold)
    stroke_width = block_stroke_width(block)
    wrapped, font, width, height = fit_text(
        draw=draw,
        text=text,
        font_path=font_path,
        font_size=block.font_size,
        max_width=block.max_width,
        max_height=block.max_height,
    )
    left = block.x - width / 2
    top = block.y - height / 2

    draw.multiline_text(
        (left, top),
        wrapped,
        fill=block.fill,
        font=font,
        align=block.align,
        spacing=8,
        stroke_width=stroke_width,
        stroke_fill=block.stroke_fill or block.fill,
    )


def safe_filename(value: str, fallback: str) -> str:
    cleaned = re.sub(r"[^\w\s.-]+", "", value, flags=re.UNICODE)
    cleaned = re.sub(r"\s+", "_", cleaned).strip("._ ")
    return cleaned[:80] or fallback


def output_extension(template_path: Path, forced_format: str | None) -> str:
    if forced_format:
        return forced_format.lower().lstrip(".")
    suffix = template_path.suffix.lower().lstrip(".")
    return "jpg" if suffix == "jpeg" else suffix or "png"


def save_image(image: Image.Image, path: Path, extension: str) -> None:
    if extension in {"jpg", "jpeg"} and image.mode != "RGB":
        background = Image.new("RGB", image.size, "white")
        if image.mode == "RGBA":
            background.paste(image, mask=image.getchannel("A"))
            image = background
        else:
            image = image.convert("RGB")
    image.save(path)


def flatten_to_rgb(image: Image.Image, background: str = "white") -> Image.Image:
    if image.mode == "RGB":
        return image
    rgb = Image.new("RGB", image.size, background)
    if image.mode == "RGBA":
        rgb.paste(image, mask=image.getchannel("A"))
    else:
        rgb.paste(image.convert("RGB"))
    return rgb


def fit_image_on_page(image: Image.Image, page_size: tuple[int, int], margin: int = 0) -> Image.Image:
    page_width, page_height = page_size
    content_width = page_width - margin * 2
    content_height = page_height - margin * 2
    scale = min(content_width / image.width, content_height / image.height)
    target_size = (max(1, int(image.width * scale)), max(1, int(image.height * scale)))

    page = Image.new("RGB", page_size, "white")
    resized = flatten_to_rgb(image).resize(target_size, Image.Resampling.LANCZOS)
    left = (page_width - target_size[0]) // 2
    top = (page_height - target_size[1]) // 2
    page.paste(resized, (left, top))
    return page


def save_pdf_pages(pages: list[Image.Image], output_path: Path) -> None:
    if not pages:
        raise ValueError("Нет страниц для PDF")
    output_path.parent.mkdir(parents=True, exist_ok=True)
    first, *rest = [flatten_to_rgb(page) for page in pages]
    first.save(
        output_path,
        "PDF",
        save_all=True,
        append_images=rest,
        resolution=PDF_DPI,
    )


def render_badge_image(row: FieldMap, sides: list[BadgeSide], template_path: Path, side_gap: int) -> Image.Image:
    rendered_sides: list[Image.Image] = []
    for side in sides:
        with Image.open(template_path) as template:
            side_image = template.convert("RGBA")
        for block in side.blocks:
            draw_centered_text(side_image, block, row.get(block.field, ""))
        rendered_sides.append(side_image)

    width = sum(image.width for image in rendered_sides) + side_gap * (len(rendered_sides) - 1)
    height = max(image.height for image in rendered_sides)
    image = Image.new("RGBA", (width, height), (255, 255, 255, 0))

    left = 0
    for side_image in rendered_sides:
        image.paste(side_image, (left, 0))
        left += side_image.width + side_gap
    return image


def best_badge_sheet_layout(badge_size: tuple[int, int]) -> dict[str, float | int | tuple[int, int]]:
    badge_width, badge_height = badge_size
    best: dict[str, float | int | tuple[int, int]] | None = None

    for page_size in (A4_PORTRAIT, A4_LANDSCAPE):
        page_width, page_height = page_size
        content_width = page_width - BADGE_PDF_MARGIN * 2
        content_height = page_height - BADGE_PDF_MARGIN * 2
        for cols in range(1, 7):
            for rows in range(1, 11):
                cell_width = (content_width - BADGE_PDF_GAP * (cols - 1)) / cols
                cell_height = (content_height - BADGE_PDF_GAP * (rows - 1)) / rows
                if cell_width <= 0 or cell_height <= 0:
                    continue
                scale = min(cell_width / badge_width, cell_height / badge_height, 1.0)
                if scale < BADGE_MIN_SCALE:
                    continue

                per_page = cols * rows
                candidate: dict[str, float | int | tuple[int, int]] = {
                    "page_size": page_size,
                    "cols": cols,
                    "rows": rows,
                    "scale": scale,
                    "per_page": per_page,
                }
                if best is None:
                    best = candidate
                    continue
                if per_page > int(best["per_page"]):
                    best = candidate
                    continue
                if per_page == int(best["per_page"]) and scale > float(best["scale"]):
                    best = candidate

    if best is None:
        page_size = A4_LANDSCAPE if badge_width >= badge_height else A4_PORTRAIT
        content_width = page_size[0] - BADGE_PDF_MARGIN * 2
        content_height = page_size[1] - BADGE_PDF_MARGIN * 2
        best = {
            "page_size": page_size,
            "cols": 1,
            "rows": 1,
            "scale": min(content_width / badge_width, content_height / badge_height),
            "per_page": 1,
        }
    return best


def make_badge_pdf_pages(badges: list[Image.Image]) -> list[Image.Image]:
    if not badges:
        return []

    layout = best_badge_sheet_layout(badges[0].size)
    page_size = layout["page_size"]
    assert isinstance(page_size, tuple)
    cols = int(layout["cols"])
    rows = int(layout["rows"])
    scale = float(layout["scale"])
    per_page = int(layout["per_page"])

    page_width, page_height = page_size
    badge_width = int(badges[0].width * scale)
    badge_height = int(badges[0].height * scale)
    grid_width = cols * badge_width + (cols - 1) * BADGE_PDF_GAP
    grid_height = rows * badge_height + (rows - 1) * BADGE_PDF_GAP
    start_x = (page_width - grid_width) // 2
    start_y = (page_height - grid_height) // 2

    pages: list[Image.Image] = []
    for page_start in range(0, len(badges), per_page):
        page = Image.new("RGB", page_size, "white")
        for slot, badge in enumerate(badges[page_start : page_start + per_page]):
            col = slot % cols
            row = slot // cols
            x = start_x + col * (badge_width + BADGE_PDF_GAP)
            y = start_y + row * (badge_height + BADGE_PDF_GAP)
            resized = flatten_to_rgb(badge).resize((badge_width, badge_height), Image.Resampling.LANCZOS)
            page.paste(resized, (x, y))
        pages.append(page)
    return pages


def render_badges(
    excel_path: Path,
    template_path: Path,
    output_dir: Path,
    config: dict[str, Any],
    sheet_name: str | None,
    forced_format: str | None,
) -> int:
    sides = build_badge_sides(config)
    fields = sorted({block.field for side in sides for block in side.blocks})
    rows = read_excel_rows(excel_path, fields, sheet_name)
    output_dir.mkdir(parents=True, exist_ok=True)

    extension = output_extension(template_path, forced_format)
    prefix = str(config["badges"].get("output_prefix", "badge"))
    side_gap = int(config["badges"].get("side_gap", 0))
    rendered_badges: list[Image.Image] = []

    for number, row in enumerate(rows, start=1):
        image = render_badge_image(row, sides, template_path, side_gap)
        if extension == "pdf":
            rendered_badges.append(image)
            continue
        name_source = row.get("full_name_ru") or row.get("full_name_en") or str(number)
        filename = f"{number:03d}_{safe_filename(name_source, prefix)}.{extension}"
        save_image(image, output_dir / filename, extension)

    if extension == "pdf":
        pages = make_badge_pdf_pages(rendered_badges)
        save_pdf_pages(pages, output_dir / "badges_a4.pdf")

    return len(rows)


def render_certificate_image(row: FieldMap, blocks: list[TextBlock], template_path: Path) -> Image.Image:
    with Image.open(template_path) as template:
        image = template.convert("RGBA")

    for block in blocks:
        draw_centered_text(image, block, row.get(block.field, ""))
    return image


def render_documents(
    mode: str,
    excel_path: Path,
    template_path: Path,
    output_dir: Path,
    config_path: Path | None,
    sheet_name: str | None,
    forced_format: str | None,
) -> int:
    config = load_json_config(config_path)

    if mode == "badges":
        return render_badges(
            excel_path=excel_path,
            template_path=template_path,
            output_dir=output_dir,
            config=config,
            sheet_name=sheet_name,
            forced_format=forced_format,
        )

    blocks = build_text_blocks(config, mode)
    fields = sorted({block.field for block in blocks})
    rows = read_excel_rows(excel_path, fields, sheet_name)
    output_dir.mkdir(parents=True, exist_ok=True)

    extension = output_extension(template_path, forced_format)
    prefix = str(config[mode].get("output_prefix", mode))
    pdf_pages: list[Image.Image] = []

    for number, row in enumerate(rows, start=1):
        image = render_certificate_image(row, blocks, template_path)

        if extension == "pdf":
            page_size = A4_LANDSCAPE if image.width >= image.height else A4_PORTRAIT
            pdf_pages.append(fit_image_on_page(image, page_size))
            continue
        name_source = (
            row.get("full_name_ru")
            or row.get("full_name_en")
            or row.get("talk_title")
            or str(number)
        )
        filename = f"{number:03d}_{safe_filename(name_source, prefix)}.{extension}"
        save_image(image, output_dir / filename, extension)

    if extension == "pdf":
        save_pdf_pages(pdf_pages, output_dir / "certificates_a4.pdf")

    return len(rows)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Генератор бейджей и сертификатов из Excel по PNG/JPG шаблону."
    )
    parser.add_argument(
        "mode",
        choices=["badges", "certificates"],
        help="Что генерировать: badges — бейджи, certificates — сертификаты.",
    )
    parser.add_argument("--excel", required=True, type=Path, help="Путь к .xlsx файлу.")
    parser.add_argument("--template", required=True, type=Path, help="Путь к PNG/JPG шаблону.")
    parser.add_argument("--output", required=True, type=Path, help="Папка для готовых файлов.")
    parser.add_argument("--config", type=Path, help="JSON с координатами текста.")
    parser.add_argument("--sheet", help="Название листа Excel. По умолчанию первый лист.")
    parser.add_argument("--format", choices=["png", "jpg", "jpeg", "pdf"], help="Формат результата.")
    return parser.parse_args()


def main() -> int:
    args = parse_args()

    try:
        count = render_documents(
            mode=args.mode,
            excel_path=args.excel,
            template_path=args.template,
            output_dir=args.output,
            config_path=args.config,
            sheet_name=args.sheet,
            forced_format=args.format,
        )
    except Exception as exc:
        print(f"Ошибка: {exc}", file=sys.stderr)
        return 1

    print(f"Готово: создано файлов — {count}. Папка: {args.output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
